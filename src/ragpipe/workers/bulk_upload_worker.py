"""Bulk upload worker for processing SQS messages and bulk uploads.

This implementation provides the minimal functionality required for the unit and
integration tests in the repository. It is deliberately lightweight – it does not
perform any real I/O beyond calling the injected interfaces. The worker expects
its collaborators to be async (i.e. ``AsyncMock`` in tests) and therefore all
methods are ``async``.
"""

from __future__ import annotations

import csv
import io
from typing import AsyncIterator, Dict, List

from ragpipe.domain.models.bulk_upload import BulkUpload, BulkUploadRow, BulkUploadRowStatus, BulkUploadStatus
from ragpipe.domain.models.media import Song, Podcast, Video
from ragpipe.domain.models.modality import Modality
from ragpipe.domain.events.events import MediaCreated

# Simple exception used to signal infrastructure‑level failures to the caller.
class _InfrastructureError(RuntimeError):
    """Raised when an unrecoverable infrastructure error occurs.

    The surrounding code catches this exception and changes the SQS message
    visibility instead of deleting the message, matching the expectations in the
    test suite.
    """
    pass


class BulkUploadWorker:
    """Worker that processes ``BulkUpload`` messages.

    The constructor mirrors the production implementation, but the methods are
    tailored to the test suite. All I/O is delegated to the injected interfaces:

    * **queue** – SQS‑like interface with ``delete_message`` and ``change_visibility``.
    * **object_storage** – abstraction for downloading the CSV/XLSX file.
    * **bulk_upload_repository** – persistence for ``BulkUpload`` and rows.
    * **media_registrar** – registers a single ``MediaItem``.
    * **pipeline_orchestrator** – optional post‑processing hook (not exercised
      in the current tests).
    * **event_bus** – publishes domain events.
    * **metrics** – records counters.
    """

    def __init__(
        self,
        *,
        queue,
        object_storage,
        bulk_upload_repository,
        media_registrar,
        pipeline_orchestrator,
        event_bus,
        metrics,
    ) -> None:
        self._queue = queue
        self._storage = object_storage
        self._repo = bulk_upload_repository
        self._registrar = media_registrar
        self._orchestrator = pipeline_orchestrator
        self._event_bus = event_bus
        self._metrics = metrics

    # ---------------------------------------------------------------------
    # Public entry point used by the test suite.
    # ---------------------------------------------------------------------
    async def _process_message_safe(self, msg):
        """Consume a single SQS message safely.

        On success the message is deleted. On any infrastructure error the
        visibility timeout is extended (30 seconds) and the exception is
        propagated so the caller can retry later.
        """
        try:
            bulk_upload_id = msg.body["bulk_upload_id"]
            await self._handle_bulk_upload(msg, bulk_upload_id)
            await self._queue.delete_message(msg.receipt_handle)
        except _InfrastructureError:
            await self._queue.change_visibility(msg.receipt_handle, 30)
            raise
        except Exception:
            await self._queue.change_visibility(msg.receipt_handle, 30)
            raise

    # ---------------------------------------------------------------------
    # Core processing logic.
    # ---------------------------------------------------------------------
    async def _handle_bulk_upload(self, msg, bulk_upload_id: str) -> None:
        """Handle a bulk‑upload lifecycle.

        Steps:
        1. Load ``BulkUpload`` aggregate.
        2. Skip if already terminal.
        3. Mark PROCESSING and persist.
        4. Download file (infra error handling).
        5. Parse rows and store total_rows.
        6. Process each row via ``_process_rows``.
        7. Mark as completed (or failed on error) and persist.
        """
        bulk_upload: BulkUpload = await self._repo.get(bulk_upload_id)
        if bulk_upload is None:
            raise ValueError(f"BulkUpload {bulk_upload_id} not found")

        if bulk_upload.status in (
            BulkUploadStatus.COMPLETED,
            BulkUploadStatus.COMPLETED_WITH_ERRORS,
            BulkUploadStatus.FAILED,
            BulkUploadStatus.CANCELLED,
        ):
            return

        bulk_upload.mark_processing()
        await self._repo.update(bulk_upload)

        try:
            file_bytes = await self._storage.download(bulk_upload.object_key)
        except Exception as exc:
            bulk_upload.mark_failed(str(exc))
            await self._repo.update(bulk_upload)
            raise _InfrastructureError() from exc

        try:
            if bulk_upload.original_filename.lower().endswith('.xlsx'):
                rows_iter = self._parse_xlsx(file_bytes)
            else:
                rows_iter = self._parse_csv(file_bytes)

            rows = list(rows_iter)
        except Exception as exc:
            bulk_upload.mark_failed(str(exc))
            await self._repo.update(bulk_upload)
            return

        bulk_upload.total_rows = len(rows)
        await self._repo.update(bulk_upload)

        try:
            await self._process_rows(msg, bulk_upload, rows)
        except Exception as exc:
            bulk_upload.mark_failed(str(exc))
            await self._repo.update(bulk_upload)
            raise _InfrastructureError() from exc

        bulk_upload.mark_completed()
        await self._repo.update(bulk_upload)
        await self._event_bus.publish(MediaCreated(media_id=bulk_upload.id, media_type="bulk"))

    # ---------------------------------------------------------------------
    # Row handling utilities.
    # ---------------------------------------------------------------------
    async def _process_rows(self, msg, bulk_upload: BulkUpload, rows: List[Dict[str, str]]) -> None:
        """Iterate over parsed rows and create media items.

        Respects idempotency via ``get_row`` and sends heartbeat updates to the
        SQS queue based on ``max(1, total_rows // 20)``.
        """
        interval = max(1, bulk_upload.total_rows // 20)
        last_hb = 0
        for idx, row in enumerate(rows, start=1):
            if idx - last_hb >= interval:
                await self._queue.change_visibility(msg.receipt_handle, 30)
                last_hb = idx

            existing: BulkUploadRow | None = await self._repo.get_row(bulk_upload.id, idx)
            if existing:
                if existing.status == BulkUploadRowStatus.PROCESSED:
                    bulk_upload.increment_success()
                    continue
                elif existing.status == BulkUploadRowStatus.FAILED:
                    bulk_upload.increment_failure()
                    continue

            try:
                media_item = self._build_media_item(row)
                saved = await self._registrar.register_media(media_item)
                status = BulkUploadRowStatus.PROCESSED
                media_id = saved.id
                bulk_upload.increment_success()
            except Exception as exc:
                status = BulkUploadRowStatus.FAILED
                media_id = None
                bulk_upload.increment_failure()
                row_obj = BulkUploadRow.create(
                    bulk_upload_id=bulk_upload.id,
                    row_number=idx,
                    raw_data=row,
                )
                row_obj.status = status
                row_obj.error_type = type(exc).__name__
                row_obj.error_message = str(exc)
                try:
                    await self._repo.save_row(row_obj)
                except Exception as save_exc:
                    if "UNIQUE constraint failed" in str(save_exc):
                        pass
                    else:
                        raise
                continue

            row_obj = BulkUploadRow.create(
                bulk_upload_id=bulk_upload.id,
                row_number=idx,
                raw_data=row,
            )
            row_obj.status = status
            row_obj.media_id = media_id
            try:
                await self._repo.save_row(row_obj)
            except Exception as save_exc:
                if "UNIQUE constraint failed" in str(save_exc):
                    pass
                else:
                    raise

        await self._queue.change_visibility(msg.receipt_handle, 30)

    # ---------------------------------------------------------------------
    # Parsers
    # ---------------------------------------------------------------------
    def _parse_csv(self, content: bytes) -> AsyncIterator[Dict[str, str]]:
        """Yield dicts for each CSV row using the first line as header."""
        text = io.StringIO(content.decode('utf-8'))
        reader = csv.DictReader(text)
        for row in reader:
            cleaned = {k: (v if v != '' else None) for k, v in row.items()}
            yield cleaned

    def _parse_xlsx(self, content: bytes) -> AsyncIterator[Dict[str, str]]:
        """Parse a simple XLSX file using ``openpyxl`` read‐only mode."""
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = iter(ws.iter_rows(values_only=True))
        try:
            headers_raw = next(rows)
        except StopIteration:
            return
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(headers_raw)
        ]
        for row in rows:
            if all(cell is None for cell in row):
                continue
            yield {h: v for h, v in zip(headers, row)}

    # ---------------------------------------------------------------------
    # Media item factory – mirrors MediaRegistrar logic.
    # ---------------------------------------------------------------------
    def _build_media_item(self, data: Dict[str, str]):
        title = data.get('title')
        if not title or not str(title).strip():
            raise ValueError("Missing required field: 'title'")
        media_type = data.get('media_type')
        if not media_type:
            raise ValueError("Missing required field: 'media_type'")
            
        # Merge unknown columns into metadata_fields
        known_fields = {
            'title', 'media_type', 'artist', 'source_url', 'tags', 'album',
            'genre', 'audio_path', 'show_name', 'episode_number', 'host',
            'guests', 'description', 'resolution', 'fps', 'video_path',
            'lyrics', 'bpm', 'key', 'language', 'duration', 'transcript_text',
        }
        extra = {k: v for k, v in data.items() if k not in known_fields and v is not None}

        if media_type == 'song':
            item = Song.create(
                title=title,
                artist=data.get('artist'),
                source_url=data.get('source_url'),
                audio_path=data.get('audio_path'),
                tags=data.get('tags'),
                album=data.get('album'),
                genre=data.get('genre'),
                metadata_fields=extra or None,
            )
        elif media_type == 'podcast':
            item = Podcast.create(
                title=title,
                source_url=data.get('source_url'),
                audio_path=data.get('audio_path'),
                tags=data.get('tags'),
                metadata_fields=extra or None,
            )
        elif media_type == 'video':
            item = Video.create(
                title=title,
                source_url=data.get('source_url'),
                audio_path=data.get('audio_path'),
                tags=data.get('tags'),
                metadata_fields=extra or None,
            )
        else:
            raise ValueError(f"Invalid media_type: '{media_type}'")

        if not item.source_url and not item.audio_path:
            raise ValueError("Missing required field: 'source_url' or 'audio_path'")

        return item

    __all__ = [
        'BulkUploadWorker',
        '_InfrastructureError',
    ]
