import pytest
from ragpipe.workers.bulk_upload_worker import BulkUploadWorker
from unittest.mock import AsyncMock
import io
import openpyxl

@pytest.fixture
def worker():
    return BulkUploadWorker(
        queue=AsyncMock(),
        object_storage=AsyncMock(),
        bulk_upload_repository=AsyncMock(),
        media_registrar=AsyncMock(),
        pipeline_orchestrator=AsyncMock(),
        event_bus=AsyncMock(),
        metrics=AsyncMock(),
    )

def create_xlsx_bytes(data, sheet_names=None):
    wb = openpyxl.Workbook()
    if sheet_names:
        ws = wb.active
        ws.title = sheet_names[0]
        for name in sheet_names[1:]:
            wb.create_sheet(name)
    else:
        ws = wb.active

    for row in data:
        ws.append(row)
    
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def test_valid_xlsx(worker):
    data = [
        ["title", "media_type", "source_url"],
        ["Song A", "song", "http://a"],
        ["Song B", "song", "http://b"],
    ]
    xlsx_bytes = create_xlsx_bytes(data)
    rows = list(worker._parse_xlsx(xlsx_bytes))
    assert len(rows) == 2
    assert rows[0]["title"] == "Song A"
    assert rows[1]["title"] == "Song B"

def test_multiple_sheets_only_active_read(worker):
    data1 = [["title"], ["A"]]
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(data1[0])
    ws1.append(data1[1])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["title"])
    ws2.append(["B"])
    
    out = io.BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()
    
    rows = list(worker._parse_xlsx(xlsx_bytes))
    assert len(rows) == 1
    assert rows[0]["title"] == "A"

def test_empty_sheet(worker):
    xlsx_bytes = create_xlsx_bytes([])
    rows = list(worker._parse_xlsx(xlsx_bytes))
    assert len(rows) == 0

def test_header_only(worker):
    xlsx_bytes = create_xlsx_bytes([["title", "media_type"]])
    rows = list(worker._parse_xlsx(xlsx_bytes))
    assert len(rows) == 0

def test_missing_columns_handled(worker):
    data = [
        ["title", None, "source_url"],
        ["A", "song", "http://a"]
    ]
    xlsx_bytes = create_xlsx_bytes(data)
    rows = list(worker._parse_xlsx(xlsx_bytes))
    assert len(rows) == 1
    assert rows[0]["title"] == "A"
    assert rows[0]["col_1"] == "song"

def test_malformed_xlsx(worker):
    with pytest.raises(Exception):
        list(worker._parse_xlsx(b"not an xlsx file"))

